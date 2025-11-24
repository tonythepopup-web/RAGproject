import os
import re
import heapq
from itertools import product
from openpyxl import load_workbook
from collections import defaultdict
import numpy as np

# ─────────────────────────────────────────────────────────────────────────────
# 🔧 텍스트 정제 및 정답 매칭
def clean_text(text):
    """문자열에서 괄호, 공백, 탭 등을 제거합니다."""
    s = str(text or "")
    return re.sub(r"[()\u00A0 \t]", "", s).strip()

def is_valid_match(candidate, answer):
    """
    후보 문장(candidate)이 정답(answer)을 포함하는지 확인합니다.
    '의' + 한 자리 숫자 형태의 매칭은 제외합니다.
    """
    pos = candidate.find(answer)
    if pos == -1:
        return False
    after = candidate[pos + len(answer):pos + len(answer) + 2]
    if after.startswith("의") and len(after) >= 2 and after[1].isdigit():
        return False
    return True

# ─────────────────────────────────────────────────────────────────────────────
# 📥 입력 처리 함수
def extract_candidate_rankings(filepath):
    """
    엑셀 파일에서 각 질문별 Top-5 후보 문장을 추출합니다.
    2행부터 5개씩 건너뛰며 E열의 값을 읽고, 
    D·E열 모두 비어 있으면 종료합니다.
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]
    last_row = ws.max_row

    rankings = []
    for base_row in range(2, last_row + 1, 5):
        if ws.cell(row=base_row, column=4).value is None and ws.cell(row=base_row, column=5).value is None:
            break
        row = [clean_text(ws.cell(row=base_row + i, column=5).value) for i in range(5)]
        rankings.append(row)

    return rankings

def extract_ground_truths(filepath):
    """
    엑셀 파일에서 각 질문별 정답 문장(Truth)을 추출합니다.
    2행부터 5개씩 건너뛰며 D열의 값을 읽고, 
    D열이 비어 있으면 종료합니다.
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]
    last_row = ws.max_row

    truths = []
    for base_row in range(2, last_row + 1, 5):
        if ws.cell(row=base_row, column=4).value is None:
            break
        gt = []
        for i in range(5):
            txt = clean_text(ws.cell(row=base_row + i, column=4).value)
            if txt:
                gt.append(txt)
        truths.append(gt)

    return truths

# ─────────────────────────────────────────────────────────────────────────────
# 📊 평가 함수
def evaluate_top5_fraction_matched(predictions, ground_truths):
    """
    Top-5 내에서 정답 포함 비율을 질문별로 계산하고 합산합니다.
    반환값: (score_adjusted, raw_total_score)
    """
    total_score = 0.0
    valid_qs = 0

    for preds, golds in zip(predictions, ground_truths):
        if not golds:
            continue
        valid_qs += 1
        match_count = 0
        for g in golds:
            for p in preds[:5]:
                if is_valid_match(p, g):
                    match_count += 1
                    break
        total_score += match_count / len(golds)

    if valid_qs == 0:
        return 0.0, 0.0
    return total_score / valid_qs, total_score

def evaluate_global_fraction(predictions, ground_truths):
    """
    전체 정답 대비 매칭된 정답 수 비율을 계산합니다.
    반환값: (fraction, matched_count, total_gold_count)
    """
    total_matched = 0
    total_gold = 0

    for preds, golds in zip(predictions, ground_truths):
        for g in golds:
            if any(is_valid_match(p, g) for p in preds[:5]):
                total_matched += 1
        total_gold += len(golds)

    if total_gold == 0:
        return 0.0, 0, 0
    total_matched -= 1  # 보정
    return total_matched / total_gold, total_matched, total_gold

def evaluate_soft(top2_list, ground_truths):
    """
    Soft Top-2 평가 방식: 정답이 하나면 1점, 둘 이상이면 match/2로 계산,
    총점은 질문 수(100)로 나눕니다.
    """
    total_score = 0
    for preds, golds in zip(top2_list, ground_truths):
        match = sum(1 for g in golds for p in preds if is_valid_match(p, g))
        if not golds:
            continue
        total_score += 1 if len(golds) == 1 else min(match / 2, 1.0)
    return total_score / 100, total_score


def evaluate_soft_global_top2(top2_list, ground_truths):
    total_matched = 0
    total_possible = 0
    for preds, golds in zip(top2_list, ground_truths):
        max_golds = min(len(golds), len(preds))
        total_possible += max_golds
        matched = sum(1 for p in preds[:2] if any(is_valid_match(p, g) for g in golds))
        total_matched += min(matched, max_golds)
    if total_possible == 0:
        return 0.0, 0, 0
    total_matched -= 1  # 보정
    return total_matched / total_possible, total_matched, total_possible



def evaluate_soft3(top3_list, ground_truths):
    """
    질문별 점수 = (Top-3로 맞춘 '서로 다른' 정답 수) / min(len(golds), 3)
    반환: (질문 평균 점수, 질문 점수 합, 유효 질문 수)
    """
    total_score = 0.0
    valid_qs = 0

    for preds, golds in zip(top3_list, ground_truths):
        if not golds:
            continue
        valid_qs += 1

        preds3 = preds[:3]
        matched_golds = set()
        for g in golds:
            if any(is_valid_match(p, g) for p in preds3):
                matched_golds.add(g)

        denom = min(len(golds), 3)
        total_score += (len(matched_golds) / denom) if denom > 0 else 0.0

    if valid_qs == 0:
        return 0.0, 0.0, 0
    return total_score / valid_qs, total_score

def evaluate_soft_global_top3(top3_list, ground_truths):
    total_matched = 0
    total_possible = 0
    for preds, golds in zip(top3_list, ground_truths):
        max_golds = min(len(golds), len(preds))
        total_possible += max_golds
        matched = sum(1 for p in preds[:3] if any(is_valid_match(p, g) for g in golds))
        total_matched += min(matched, max_golds)
    if total_possible == 0:
        return 0.0, 0, 0
    total_matched -= 1  # 보정
    return total_matched  / total_possible, total_matched, total_possible

def evaluate_soft4(top4_list, ground_truths):
    """
    질문별 점수 = (Top-4로 맞춘 '서로 다른' 정답 수) / min(len(golds), 4)
    반환: (질문 평균 점수, 질문 점수 합, 유효 질문 수)
    """
    total_score = 0.0
    valid_qs = 0

    for preds, golds in zip(top4_list, ground_truths):
        if not golds:
            continue
        valid_qs += 1

        preds4 = preds[:4]
        matched_golds = set()
        for g in golds:
            if any(is_valid_match(p, g) for p in preds4):
                matched_golds.add(g)

        denom = min(len(golds), 4)
        total_score += (len(matched_golds) / denom) if denom > 0 else 0.0

    if valid_qs == 0:
        return 0.0, 0.0, 0
    return total_score  / valid_qs, total_score


def evaluate_soft_global_top4(top4_list, ground_truths):
    total_matched = 0
    total_possible = 0
    for preds, golds in zip(top4_list, ground_truths):
        max_golds = min(len(golds), len(preds))
        total_possible += max_golds
        matched = sum(1 for p in preds[:4] if any(is_valid_match(p, g) for g in golds))
        total_matched += min(matched, max_golds)
    if total_possible == 0:
        return 0.0, 0, 0
    total_matched -= 1  # 보정
    return total_matched / total_possible, total_matched, total_possible

# ─────────────────────────────────────────────────────────────────────────────
# 🤖 앙상블 함수
def weighted_ensemble_rankings_top5(models, weights):
    """가중치 앙상블로 Top-5 후보를 생성합니다."""
    ensemble_top5 = []
    for q in range(len(models[0])):
        score = defaultdict(float)
        for idx, model in enumerate(models):
            for rank, cand in enumerate(model[q]):
                if cand:
                    score[cand] += (5 - rank) * weights[idx]
        sorted_cands = sorted(score.items(), key=lambda x: -x[1])
        ensemble_top5.append([c for c, _ in sorted_cands[:5]])
    return ensemble_top5

def weighted_ensemble_rankings_top2(models, weights):
    """가중치 앙상블로 Top-2 후보를 생성합니다."""
    ensemble_top2 = []
    for q in range(len(models[0])):
        score = defaultdict(float)
        for idx, model in enumerate(models):
            for rank, cand in enumerate(model[q]):
                if cand:
                    score[cand] += (5 - rank) * weights[idx]
        sorted_cands = sorted(score.items(), key=lambda x: -x[1])
        ensemble_top2.append([c for c, _ in sorted_cands[:2]])
    return ensemble_top2

def run_weight_grid_search_v2_light(
    models,
    model_names,
    ground_truths,
    weight_range=(0.5, 1.5),
    step=0.25,
    normalize=True,
    top_n=5
):
    weight_candidates = np.arange(weight_range[0], weight_range[1] + step, step)
    combos = list(product(weight_candidates, repeat=len(models)))
    best_heap = []

    print(f"\n🔬 총 {len(combos):,}개의 가중치 조합 실험 중...")

    for weights in combos:
        if normalize:
            s = sum(weights)
            if s == 0:
                continue
            weights = [w / s for w in weights]

        ensemble_top2 = weighted_ensemble_rankings_top2(models, weights)
        frac2, matched2, possible2 = evaluate_soft_global_top2(ensemble_top2, ground_truths)
        heapq.heappush(best_heap, (frac2, matched2, possible2, list(weights)))
        
        if len(best_heap) > top_n:
            heapq.heappop(best_heap)

    best_heap.sort(reverse=True)
    print(f"\n✅ 앙상블 그리드 탐색 Top-{top_n} 결과:")
    for i, result in enumerate(best_heap, 1):
        if len(result) == 4:
            frac, matched, total, weights = result
            print(f"\n🥇 Rank {i} | 정답포함률 = {frac:.4f} ({matched}/{total})")
        else:
            frac, matched, weights = result
            print(f"\n🥇 Rank {i} | 정확도 = {frac:.4f} ({matched:.1f}/100)")
        
        for name, w in zip(model_names, weights):
            print(f" - {name}: {w:.3f}")
    if len(best_heap[0]) == 4:
        return best_heap[0][3], best_heap[0][0]  # weights, score
    else:
        return best_heap[0][2], best_heap[0][0]

# ─────────────────────────────────────────────────────────────────────────────
# 🧪 실행 파트
if __name__ == "__main__":
    from pathlib import Path
    SCRIPT_DIR = Path(__file__).resolve().parent
    model_folder = str(SCRIPT_DIR.parent.parent / "00_data" / "output" / "benchmark_result")
    model_weights_dict = {
        # "cosine_idf_0.7_벤치마크v2_0711_전처리완료_라우팅포함": 1,
        # "cosine_idf_0.8_벤치마크v2_0711_전처리완료_라우팅포함": 1,
        "cosine_idf_0.9_벤치마크v2_0711_전처리완료_라우팅포함": 1,
        "벤치마크v2_0711_전처리완료_라우팅포함" : 1, 
        "keyword_벤치마크v2_0711_전처리완료_라우팅포함" : 1,
        "rocchio feedback_벤치마크v2_0711_전처리완료_라우팅포함" : 1,
        # "오직Keyword_벤치마크v2_0711_전처리완료_라우팅포함" : 1,
        # "형태소분석_0502_벤치마크v2_0711_전처리완료_라우팅포함" : 1,
        "형태소분석_07005_벤치마크v2_0711_전처리완료_라우팅포함": 1

    }

    # 모델 파일 로드
    model_files_all = sorted( 
        f for f in os.listdir(model_folder) if f.endswith(".xlsm")
    )
    model_files_all = [os.path.join(model_folder, f) for f in model_files_all]
    model_names_all = [os.path.splitext(os.path.basename(f))[0] for f in model_files_all]

    selected_files, selected_weights, selected_names = [], [], []
    print("🔍 사용된 모델 목록:")
    for name, path in zip(model_names_all, model_files_all):
        if name in model_weights_dict:
            selected_files.append(path)
            selected_weights.append(model_weights_dict[name])
            selected_names.append(name)
            print(f" - {name} (w={model_weights_dict[name]:.2f})")

    print("\n📥 모델 불러오는 중...")
    ground_truths = extract_ground_truths(selected_files[0])
    model_rankings = [extract_candidate_rankings(f) for f in selected_files]
    all_model_rankings = [extract_candidate_rankings(f) for f in model_files_all]

   

    # 3️⃣ 전체 모델 Top-2 평가
    print("\n📊 전체 모델 Top-2 정답 포함률:")
    for name, ranking in zip(model_names_all, all_model_rankings):
        top2_list = [r[:2] for r in ranking]
        acc2, score2 = evaluate_soft(top2_list, ground_truths)
        print(f"{name}: 정확도 = {acc2:.3f} ({score2:.1f}/100)")

    # 4️⃣ 전체 모델 Top-2 Soft 전역 평가
    print("\n📊 전체 모델 Top-2 정답 포함률 (전역 평가):")
    for name, ranking in zip(model_names_all, all_model_rankings):
        top2_list = [r[:2] for r in ranking]
        soft_frac, soft_matched, soft_total = evaluate_soft_global_top2(top2_list, ground_truths)
        print(f"{name}: 정답포함률 = {soft_frac:.3f} ({soft_matched}/{soft_total})")
        
        
    print("\n📊 전체 모델 Top-3 정답 포함률:")
    for name, ranking in zip(model_names_all, all_model_rankings):
        top3_list = [r[:3] for r in ranking]
        acc2, score2 = evaluate_soft3(top3_list, ground_truths)
        print(f"{name}: 정확도 = {acc2:.3f} ({score2:.1f}/100)")

    print("\n📊 전체 모델 Top-3 정답 포함률 (전역 평가):")
    for name, ranking in zip(model_names_all, all_model_rankings):
        top3_list = [r[:3] for r in ranking]
        soft_frac, soft_matched, soft_total = evaluate_soft_global_top3(top3_list, ground_truths)
        print(f"{name}: 정답포함률 = {soft_frac:.3f} ({soft_matched}/{soft_total})")
        
        
    print("\n📊 전체 모델 Top-4 정답 포함률:")
    for name, ranking in zip(model_names_all, all_model_rankings):
        top4_list = [r[:4] for r in ranking]
        acc2, score2 = evaluate_soft4(top4_list, ground_truths)
        print(f"{name}: 정확도 = {acc2:.3f} ({score2:.1f}/100)")

    print("\n📊 전체 모델 Top- 정답 포함률 (전역 평가):")
    for name, ranking in zip(model_names_all, all_model_rankings):
        top4_list = [r[:4] for r in ranking]
        soft_frac, soft_matched, soft_total = evaluate_soft_global_top4(top4_list, ground_truths)
        print(f"{name}: 정답포함률 = {soft_frac:.3f} ({soft_matched}/{soft_total})")
            
        
     # 1️⃣ 전체 모델 Top-5 평가
    print("\n📊 전체 모델 Top-5 정답 포함률:")
    for name, ranking in zip(model_names_all, all_model_rankings):
        acc, score = evaluate_top5_fraction_matched(ranking, ground_truths)
        print(f"{name}: 정확도 = {acc:.3f} ({score:.1f}/100)")

    # 2️⃣ 전체 모델 Top-5 전역 평가
    print("\n📊 전체 모델 Top-5 정답 포함률 (전역 평가):")
    for name, ranking in zip(model_names_all, all_model_rankings):
        acc, matched, total = evaluate_global_fraction(ranking, ground_truths)
        print(f"{name}: 정답포함률 = {acc:.3f} ({matched}/{total})")

    # 5️⃣ 앙상블 그리드 탐색 실행
    # run_weight_grid_search_v2_light(
    #     models=model_rankings,
    #     model_names=selected_names,
    #     ground_truths=ground_truths,
    #     weight_range=(0.5, 1.2),
    #     step=0.1,
    #     normalize=True,
    #     top_n=5,
    # )
